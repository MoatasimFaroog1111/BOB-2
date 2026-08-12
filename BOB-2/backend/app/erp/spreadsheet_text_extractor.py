"""Bounded text extraction for supported spreadsheet documents.

The document AI only needs a deterministic textual representation of workbook cells. This
component owns spreadsheet parsing so OCR/image code never has to know about Excel formats.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable


class SpreadsheetExtractionError(RuntimeError):
    pass


@dataclass(frozen=True, slots=True)
class SpreadsheetExtractionLimits:
    max_sheets: int = 50
    max_rows_per_sheet: int = 50_000
    max_columns_per_sheet: int = 256
    max_total_cells: int = 500_000
    max_text_chars: int = 2_000_000


class SpreadsheetTextExtractor:
    """Extract cell values from .xlsx and legacy .xls without executing workbook content."""

    supported_extensions = frozenset({".xlsx", ".xls"})

    def __init__(self, limits: SpreadsheetExtractionLimits | None = None) -> None:
        self._limits = limits or SpreadsheetExtractionLimits()

    def supports(self, file_path: str | Path) -> bool:
        return Path(file_path).suffix.lower() in self.supported_extensions

    def extract(self, file_path: str | Path) -> str:
        path = Path(file_path)
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            return self._extract_xlsx(path)
        if suffix == ".xls":
            return self._extract_xls(path)
        raise SpreadsheetExtractionError(f"Unsupported spreadsheet extension: {suffix}")

    @staticmethod
    def _cell_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ", timespec="seconds")
        if isinstance(value, date):
            return value.isoformat()
        return str(value).replace("\r\n", " ").replace("\r", " ").replace("\n", " ").strip()

    def _render_sheet(self, name: str, rows: Iterable[Iterable[Any]], *, total_cells: list[int]) -> str:
        lines = [f"[Sheet: {name}]"]
        text_chars = len(lines[0])

        for row_index, row in enumerate(rows):
            if row_index >= self._limits.max_rows_per_sheet:
                raise SpreadsheetExtractionError("Spreadsheet exceeds the safe row limit")

            values = list(row)
            if len(values) > self._limits.max_columns_per_sheet:
                raise SpreadsheetExtractionError("Spreadsheet exceeds the safe column limit")

            total_cells[0] += len(values)
            if total_cells[0] > self._limits.max_total_cells:
                raise SpreadsheetExtractionError("Spreadsheet exceeds the safe cell limit")

            rendered = "\t".join(self._cell_text(value) for value in values).rstrip()
            if not rendered:
                continue

            text_chars += len(rendered) + 1
            if text_chars > self._limits.max_text_chars:
                raise SpreadsheetExtractionError("Spreadsheet text exceeds the safe extraction limit")
            lines.append(rendered)

        return "\n".join(lines)

    def _extract_xlsx(self, path: Path) -> str:
        import openpyxl

        workbook = openpyxl.load_workbook(
            filename=str(path),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            if len(workbook.worksheets) > self._limits.max_sheets:
                raise SpreadsheetExtractionError("Spreadsheet exceeds the safe worksheet limit")

            total_cells = [0]
            sections = [
                self._render_sheet(
                    worksheet.title,
                    worksheet.iter_rows(values_only=True),
                    total_cells=total_cells,
                )
                for worksheet in workbook.worksheets
            ]
            return "\n\n".join(section for section in sections if section.strip()).strip()
        finally:
            workbook.close()

    def _extract_xls(self, path: Path) -> str:
        import xlrd

        try:
            workbook = xlrd.open_workbook(
                filename=str(path),
                on_demand=True,
                formatting_info=False,
            )
        except Exception as exc:
            raise SpreadsheetExtractionError("Invalid or unsupported XLS workbook") from exc

        try:
            if workbook.nsheets > self._limits.max_sheets:
                raise SpreadsheetExtractionError("Spreadsheet exceeds the safe worksheet limit")

            total_cells = [0]
            sections: list[str] = []
            for sheet_index in range(workbook.nsheets):
                worksheet = workbook.sheet_by_index(sheet_index)

                def rows() -> Iterable[list[Any]]:
                    for row_index in range(worksheet.nrows):
                        values: list[Any] = []
                        for col_index in range(worksheet.ncols):
                            cell = worksheet.cell(row_index, col_index)
                            value: Any = cell.value
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                try:
                                    value = xlrd.xldate_as_datetime(value, workbook.datemode)
                                except Exception:
                                    value = cell.value
                            values.append(value)
                        yield values

                sections.append(
                    self._render_sheet(
                        worksheet.name,
                        rows(),
                        total_cells=total_cells,
                    )
                )
                workbook.unload_sheet(sheet_index)

            return "\n\n".join(section for section in sections if section.strip()).strip()
        finally:
            workbook.release_resources()


spreadsheet_text_extractor = SpreadsheetTextExtractor()
